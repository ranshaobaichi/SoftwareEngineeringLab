"""
导出服务 - 处理数据导出为Excel
"""
import uuid
from typing import List, Optional
from datetime import datetime
from decimal import Decimal

from ..models.entry import Entry
from ..database.database import Database


class ExportService:
    """
    导出服务类 - 提供数据导出功能
    
    Attributes:
        database: 数据库实例
    """
    
    def __init__(self, database: Database):
        """
        初始化导出服务
        
        Args:
            database: 数据库实例
        """
        self.database = database
    
    def export_to_xlsx(
        self,
        entries: List[Entry],
        file_path: str,
        include_tags: bool = True,
        include_images: bool = True
    ) -> bool:
        """
        导出账目到Excel文件
        
        Args:
            entries: 要导出的账目列表
            file_path: 导出文件路径
            include_tags: 是否包含标签信息
            include_images: 是否包含图片信息
            
        Returns:
            是否导出成功
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "账目明细"
            
            # 设置表头
            headers = ['日期', '标题', '分类', '类型', '金额', '货币', '备注']
            if include_tags:
                headers.append('标签')
            if include_images:
                headers.append('附件数量')
            
            # 写入表头并设置样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入数据
            for row, entry in enumerate(entries, 2):
                ws.cell(row=row, column=1, value=entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row, column=2, value=entry.title)
                ws.cell(row=row, column=3, value=entry.category.name)
                ws.cell(row=row, column=4, value=entry.category.type.value)
                ws.cell(row=row, column=5, value=float(entry.amount))
                ws.cell(row=row, column=6, value=entry.currency)
                ws.cell(row=row, column=7, value=entry.note)
                
                col = 8
                if include_tags:
                    tags_str = ', '.join(tag.name for tag in entry.tags)
                    ws.cell(row=row, column=col, value=tags_str)
                    col += 1
                
                if include_images:
                    ws.cell(row=row, column=col, value=len(entry.images))
            
            # 调整列宽
            ws.column_dimensions['A'].width = 20  # 日期
            ws.column_dimensions['B'].width = 25  # 标题
            ws.column_dimensions['C'].width = 12  # 分类
            ws.column_dimensions['D'].width = 10  # 类型
            ws.column_dimensions['E'].width = 12  # 金额
            ws.column_dimensions['F'].width = 8   # 货币
            ws.column_dimensions['G'].width = 30  # 备注
            
            # 保存文件
            wb.save(file_path)
            return True
            
        except ImportError:
            print("错误: 需要安装 openpyxl 库来导出Excel文件")
            print("请运行: pip install openpyxl")
            return False
        except Exception as e:
            print(f"导出失败: {e}")
            return False
    
    def export_to_csv(
        self,
        entries: List[Entry],
        file_path: str,
        include_tags: bool = True
    ) -> bool:
        """
        导出账目到CSV文件
        
        Args:
            entries: 要导出的账目列表
            file_path: 导出文件路径
            include_tags: 是否包含标签信息
            
        Returns:
            是否导出成功
        """
        try:
            import csv
            
            with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                # 准备表头
                fieldnames = ['日期', '标题', '分类', '类型', '金额', '货币', '备注']
                if include_tags:
                    fieldnames.append('标签')
                
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                # 写入数据
                for entry in entries:
                    row = {
                        '日期': entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                        '标题': entry.title,
                        '分类': entry.category.name,
                        '类型': entry.category.type.value,
                        '金额': str(entry.amount),
                        '货币': entry.currency,
                        '备注': entry.note
                    }
                    
                    if include_tags:
                        row['标签'] = ', '.join(tag.name for tag in entry.tags)
                    
                    writer.writerow(row)
            
            return True
            
        except Exception as e:
            print(f"导出CSV失败: {e}")
            return False
    
    def export_statistics_to_xlsx(
        self,
        user_id: uuid.UUID,
        file_path: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> bool:
        """
        导出统计数据到Excel
        
        Args:
            user_id: 用户ID
            file_path: 导出文件路径
            start_date: 起始日期(可选)
            end_date: 结束日期(可选)
            
        Returns:
            是否导出成功
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.chart import PieChart, Reference
            
            from ..services.stat_engine import StatEngine
            from ..models.category import CategoryType
            
            stat_engine = StatEngine(self.database)
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            
            # 移除默认的工作表
            wb.remove(wb.active)
            
            # 创建汇总工作表
            ws_summary = wb.create_sheet("汇总")
            ws_summary.append(['统计项', '金额'])
            
            # 计算总收入、总支出、余额
            total_income = stat_engine.calculate_total_by_type(
                user_id, CategoryType.INCOME, start_date, end_date
            )
            total_expense = stat_engine.calculate_total_by_type(
                user_id, CategoryType.EXPENSE, start_date, end_date
            )
            balance = total_income - total_expense
            
            ws_summary.append(['总收入', float(total_income)])
            ws_summary.append(['总支出', float(total_expense)])
            ws_summary.append(['余额', float(balance)])
            
            # 创建分类统计工作表
            ws_category = wb.create_sheet("分类统计")
            ws_category.append(['分类', '金额', '次数', '百分比'])
            
            category_stats = stat_engine.get_statistics_by_category(
                user_id, start_date, end_date
            )
            
            for category_name, stats in category_stats.items():
                ws_category.append([
                    category_name,
                    float(stats['amount']),
                    stats['count'],
                    f"{stats['percentage']:.2f}%"
                ])
            
            # 保存文件
            wb.save(file_path)
            return True
            
        except ImportError:
            print("错误: 需要安装 openpyxl 库来导出Excel文件")
            return False
        except Exception as e:
            print(f"导出统计数据失败: {e}")
            return False
